"""XLSX/CSV 解析器：按工作表和行批次产生带位置的文本块。"""

import csv
from pathlib import Path
from typing import Iterable

import chardet
from openpyxl import load_workbook

from app.parsers.base import ParsedBlock


def _row_text(headers: list[str], values: Iterable[object]) -> str:
    fields: list[str] = []
    for index, value in enumerate(values):
        if value is None or str(value).strip() == "":
            continue
        label = headers[index] if index < len(headers) and headers[index] else f"列{index + 1}"
        fields.append(f"{label}={value}")
    return "；".join(fields)


class XlsxParser:
    name = "openpyxl"
    version = "1"

    def parse(self, path: Path) -> list[ParsedBlock]:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        blocks: list[ParsedBlock] = []
        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                first = next(rows, None)
                if first is None:
                    continue
                headers = [str(value).strip() if value is not None else "" for value in first]
                for row_number, values in enumerate(rows, start=2):
                    content = _row_text(headers, values)
                    if content:
                        blocks.append(ParsedBlock(content=content, sheet_name=sheet.title, row_start=row_number, row_end=row_number, section_path=[sheet.title]))
        finally:
            workbook.close()
        return blocks


class CsvParser:
    name = "csv"
    version = "1"

    def parse(self, path: Path) -> list[ParsedBlock]:
        raw = path.read_bytes()
        encoding = chardet.detect(raw).get("encoding") or "utf-8"
        text = raw.decode(encoding, errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(text.splitlines(), dialect)
        first = next(reader, None)
        if first is None:
            return []
        headers = [value.strip() for value in first]
        blocks: list[ParsedBlock] = []
        for row_number, values in enumerate(reader, start=2):
            content = _row_text(headers, values)
            if content:
                blocks.append(ParsedBlock(content=content, sheet_name="CSV", row_start=row_number, row_end=row_number, section_path=["CSV"] ))
        return blocks
